import openpyxl
from datetime import date
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q
from django.http import HttpResponse
from django.template.loader import render_to_string
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import Beneficiario, TutorPadre, PreInscripcion
from .forms import BeneficiarioForm, TutorPadreForm, PreInscripcionForm
from ninos.models import Nino

try:
    from xhtml2pdf import pisa
    PISA_AVAILABLE = True
except ImportError:
    PISA_AVAILABLE = False

BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
HEADER_FILL = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

def _es_admin_director(user):
    if not user.is_authenticated or not user.id_rol: return False
    return user.id_rol.nombre_rol in ['Administrador', 'Director', 'Directora']

def _header_cell(cell, value):
    cell.value = value
    cell.font = Font(bold=True, size=9, color='FFFFFF')
    cell.fill = HEADER_FILL
    cell.border = BORDER
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def _data_cell(cell, fill_color='FFFFFF'):
    cell.border = BORDER
    cell.font = Font(size=9)
    cell.alignment = Alignment(vertical='center')
    if fill_color != 'FFFFFF':
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

# =================== PADRES/IDS ===================
@login_required
def lista_beneficiarios(request):
    q = request.GET.get('q', '')
    beneficiarios = Beneficiario.objects.prefetch_related('ninos', 'tutores')
    if q:
        beneficiarios = beneficiarios.filter(
            Q(nombres__icontains=q) | Q(apellido_paterno__icontains=q) |
            Q(apellido_materno__icontains=q) | Q(CI_beneficiario__icontains=q)
        )
    return render(request, 'beneficiarios/lista.html', {'beneficiarios': beneficiarios, 'q': q})

@login_required
def detalle_beneficiario(request, ci):
    b = get_object_or_404(Beneficiario, CI_beneficiario=ci)
    return render(request, 'beneficiarios/detalle.html', {'beneficiario': b})

@login_required
def crear_beneficiario(request):
    form = BeneficiarioForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        b = form.save(commit=False)
        b.registrado_por = request.user
        b.save()
        messages.success(request, 'Padre IDS registrado correctamente.')
        return redirect('beneficiarios:lista_beneficiarios')
    return render(request, 'beneficiarios/form.html', {'form': form, 'titulo': 'Nuevo Padre IDS'})

@login_required
def editar_beneficiario(request, ci):
    if not _es_admin_director(request.user):
        messages.error(request, 'No tienes permiso.')
        return redirect('beneficiarios:lista_beneficiarios')
    b = get_object_or_404(Beneficiario, CI_beneficiario=ci)
    form = BeneficiarioForm(request.POST or None, request.FILES or None, instance=b)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Padre IDS actualizado.')
        return redirect('beneficiarios:lista_beneficiarios')
    return render(request, 'beneficiarios/form.html', {'form': form, 'titulo': 'Editar Padre IDS', 'objeto': b})

@login_required
def eliminar_beneficiario(request, ci):
    if not _es_admin_director(request.user):
        messages.error(request, 'No tienes permiso.')
        return redirect('beneficiarios:lista_beneficiarios')
    b = get_object_or_404(Beneficiario, CI_beneficiario=ci)
    from django.db.models.deletion import ProtectedError
    if request.method == 'POST':
        try:
            b.delete()
            messages.success(request, 'Padre IDS eliminado.')
        except ProtectedError:
            messages.error(request, 'No se puede eliminar: tiene niños registrados.')
        return redirect('beneficiarios:lista_beneficiarios')
    return render(request, 'confirmar_eliminar.html', {'objeto': b, 'tipo': 'beneficiario'})

@login_required
def exportar_excel_beneficiarios(request):
    gestion = date.today().year
    beneficiarios = Beneficiario.objects.all().order_by('apellido_paterno', 'apellido_materno', 'nombres')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Padres_IDS"

    # Colores — verde suave (ITDB) y azul suave (externos)
    AZUL_HD  = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    AMARILLO = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    VERDE_F  = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    AZUL_F   = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')

    # Fila 1 — Título (A1:L1 todos azul)
    ws.row_dimensions[1].height = 30
    for col_i in range(1, 13):
        tc = ws.cell(1, col_i)
        tc.fill = AZUL_HD
        tc.font = Font(bold=True, size=12, color='FFFFFF')
        tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(1, 1).value = f"CENTRO INFANTIL DON BOSQUITO - LISTA DE PADRES/IDS - GESTION {gestion}"

    # Fila 2 — Headers
    headers = ['Nro', 'CI', 'APELLIDO PATERNO', 'APELLIDO MATERNO', 'NOMBRES',
               'OCUPACION (ESPECIFICAR)', 'CARRERA', 'ANO/SEMESTRE', 'TURNO',
               'TELEFONO', 'DIRECCION', 'NINOS A CARGO']
    ws.append(headers)
    ws.row_dimensions[2].height = 48
    for cell in ws[2]:
        cell.fill = AMARILLO
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

    # Datos
    itdb_count = 0
    externos_count = 0
    for idx, b in enumerate(beneficiarios, 1):
        ocu = b.ocupacion or ''
        es_itdb = 'ITDB' in ocu.upper() or 'ESTUDIANTE' in ocu.upper()
        if es_itdb:
            itdb_count += 1
            row_fill = VERDE_F
        else:
            externos_count += 1
            row_fill = AZUL_F

        row_data = [
            idx,
            b.CI_beneficiario,
            (b.apellido_paterno or '').upper(),
            (b.apellido_materno or '').upper(),
            (b.nombres or '').upper(),
            b.get_ocupacion_display() if hasattr(b, 'get_ocupacion_display') else (b.ocupacion or ''),
            b.carrera or '',
            b.anio_semestre or '',
            b.get_turno_display() if hasattr(b, 'get_turno_display') else (b.turno or ''),
            b.contacto or '',
            b.direccion or '',
            b.ninos.count(),
        ]
        ws.append(row_data)
        r = ws.max_row
        ws.row_dimensions[r].height = 15
        for cell in ws[r]:
            cell.fill = row_fill
            cell.border = BORDER
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical='center')

    # Estadísticas
    ws.append([])
    ws.append([])
    est_row = ws.max_row
    ws.row_dimensions[est_row].height = 20
    for col_i in range(1, 13):
        ec = ws.cell(est_row, col_i)
        ec.fill = AZUL_HD
        ec.font = Font(bold=True, size=11, color='FFFFFF')
        ec.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(est_row, 1).value = f"ESTADISTICAS GENERALES - GESTION {gestion}"

    ws.append([])
    gh = ws.max_row
    ws.row_dimensions[gh].height = 18
    for col, val in [(1, 'CATEGORIA'), (2, 'CANTIDAD')]:
        c = ws.cell(gh, col); c.value = val
        c.font = Font(bold=True, size=9, color='FFFFFF')
        c.fill = AZUL_HD; c.border = BORDER
        c.alignment = Alignment(horizontal='center', vertical='center')

    for label, val, fill_obj in [
        ('Padres Estudiantes ITDB', itdb_count,               VERDE_F),
        ('Familias Externas',       externos_count,            AZUL_F),
        ('TOTAL PADRES/IDS',        itdb_count + externos_count,
         PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')),
    ]:
        ws.append([label, val])
        r = ws.max_row
        ws.row_dimensions[r].height = 15
        for col in [1, 2]:
            c = ws.cell(r, col)
            c.fill = fill_obj; c.border = BORDER
            c.font = Font(bold=label.startswith('TOTAL'), size=10)
            c.alignment = Alignment(horizontal='center' if col == 2 else 'left', vertical='center')

    # Anchos de columna
    for i, w in enumerate([4, 12, 18, 18, 24, 28, 22, 13, 10, 14, 28, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="PADRES_IDS_' + str(gestion) + '.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_pdf_beneficiarios(request):
    if not PISA_AVAILABLE:
        return HttpResponse("Instale xhtml2pdf.")
    gestion = date.today().year
    beneficiarios = list(Beneficiario.objects.all().order_by('apellido_paterno', 'apellido_materno', 'nombres'))
    itdb = sum(1 for b in beneficiarios if 'ITDB' in (b.ocupacion or '').upper() or 'ESTUDIANTE' in (b.ocupacion or '').upper())
    stats = {'itdb': itdb, 'externos': len(beneficiarios) - itdb, 'total': len(beneficiarios)}
    html = render_to_string('beneficiarios/reporte_pdf.html', {
        'beneficiarios': beneficiarios, 'stats': stats, 'gestion': gestion
    })
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="BENEFICIARIOS_{gestion}.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response

# =================== PADRES/EXTERNOS ===================
@login_required
def lista_tutores(request):
    q = request.GET.get('q', '')
    tutores = TutorPadre.objects.select_related('beneficiario')
    if q:
        tutores = tutores.filter(
            Q(nombres__icontains=q) | Q(apellido_paterno__icontains=q) | Q(CI_tutor__icontains=q)
        )
    return render(request, 'tutores/lista.html', {'tutores': tutores, 'q': q})

@login_required
def crear_tutor(request):
    form = TutorPadreForm(request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        t = form.save(commit=False)
        t.registrado_por = request.user
        t.save()
        messages.success(request, 'Padre Externo registrado correctamente.')
        return redirect('beneficiarios:lista_tutores')
    return render(request, 'tutores/form.html', {'form': form, 'titulo': 'Nuevo Padre Externo'})

@login_required
def editar_tutor(request, ci):
    if not _es_admin_director(request.user):
        messages.error(request, 'Sin permiso.')
        return redirect('beneficiarios:lista_tutores')
    t = get_object_or_404(TutorPadre, CI_tutor=ci)
    form = TutorPadreForm(request.POST or None, request.FILES or None, instance=t)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Padre Externo actualizado.')
        return redirect('beneficiarios:lista_tutores')
    return render(request, 'tutores/form.html', {'form': form, 'titulo': 'Editar Padre Externo', 'objeto': t})

@login_required
def detalle_tutor(request, ci):
    t = get_object_or_404(TutorPadre, CI_tutor=ci)
    return render(request, 'tutores/detalle.html', {'tutor': t})

@login_required
def eliminar_tutor(request, ci):
    if not _es_admin_director(request.user):
        messages.error(request, 'Sin permiso.')
        return redirect('beneficiarios:lista_tutores')
    t = get_object_or_404(TutorPadre, CI_tutor=ci)
    from django.db.models.deletion import ProtectedError
    if request.method == 'POST':
        try:
            t.delete()
            messages.success(request, 'Padre Externo eliminado.')
        except ProtectedError:
            messages.error(request, 'No se puede eliminar por restricciones.')
        return redirect('beneficiarios:lista_tutores')
    return render(request, 'confirmar_eliminar.html', {'objeto': t, 'tipo': 'tutor'})

@login_required
def exportar_excel_tutores(request):
    gestion = date.today().year
    tutores = TutorPadre.objects.all().order_by('apellido_paterno', 'apellido_materno', 'nombres')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Padres_Externos"

    # Colores — naranja suave (ITDB) y lavanda suave (externos)
    NARANJA_HD = PatternFill(start_color='7C3A00', end_color='7C3A00', fill_type='solid')
    AMARILLO   = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    NARANJA_F  = PatternFill(start_color='FFF7ED', end_color='FFF7ED', fill_type='solid')
    LAVANDA_F  = PatternFill(start_color='F3F0FF', end_color='F3F0FF', fill_type='solid')

    # Fila 1 — Título (A1:L1 todos naranja oscuro)
    ws.row_dimensions[1].height = 30
    for col_i in range(1, 13):
        tc = ws.cell(1, col_i)
        tc.fill = NARANJA_HD
        tc.font = Font(bold=True, size=12, color='FFFFFF')
        tc.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(1, 1).value = f"CENTRO INFANTIL DON BOSQUITO - LISTA DE PADRES/EXTERNOS - GESTION {gestion}"

    # Fila 2 — Headers
    headers = ['Nro', 'CI TUTOR', 'APELLIDO PATERNO', 'APELLIDO MATERNO', 'NOMBRES',
               'OCUPACION (ESPECIFICAR)', 'CARRERA', 'ANO/SEMESTRE', 'TURNO',
               'TELEFONO', 'DIRECCION', 'NINOS A CARGO']
    ws.append(headers)
    ws.row_dimensions[2].height = 48
    for cell in ws[2]:
        cell.fill = AMARILLO
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

    # Datos
    itdb_count = 0
    externos_count = 0
    for idx, t in enumerate(tutores, 1):
        ocu = t.ocupacion or ''
        es_itdb = 'ITDB' in ocu.upper() or 'ESTUDIANTE' in ocu.upper()
        if es_itdb:
            itdb_count += 1
            row_fill = NARANJA_F
        else:
            externos_count += 1
            row_fill = LAVANDA_F

        row_data = [
            idx, t.CI_tutor,
            (t.apellido_paterno or '').upper(),
            (t.apellido_materno or '').upper(),
            (t.nombres or '').upper(),
            t.get_ocupacion_display() if hasattr(t, 'get_ocupacion_display') else (t.ocupacion or ''),
            t.carrera or '',
            t.anio_semestre or '',
            t.get_turno_display() if hasattr(t, 'get_turno_display') else (t.turno or ''),
            t.contacto or '',
            t.direccion or '',
            t.ninos.count(),
        ]
        ws.append(row_data)
        r = ws.max_row
        ws.row_dimensions[r].height = 15
        for cell in ws[r]:
            cell.fill = row_fill
            cell.border = BORDER
            cell.font = Font(size=10)
            cell.alignment = Alignment(vertical='center')

    # Estadísticas
    ws.append([])
    ws.append([])
    est_row = ws.max_row
    ws.row_dimensions[est_row].height = 20
    for col_i in range(1, 13):
        ec = ws.cell(est_row, col_i)
        ec.fill = NARANJA_HD
        ec.font = Font(bold=True, size=11, color='FFFFFF')
        ec.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(est_row, 1).value = f"ESTADISTICAS GENERALES - GESTION {gestion}"

    ws.append([])
    gh = ws.max_row
    ws.row_dimensions[gh].height = 18
    for col, val in [(1, 'CATEGORIA'), (2, 'CANTIDAD')]:
        c = ws.cell(gh, col); c.value = val
        c.font = Font(bold=True, size=9, color='FFFFFF')
        c.fill = NARANJA_HD; c.border = BORDER
        c.alignment = Alignment(horizontal='center', vertical='center')

    for label, val, fill_obj in [
        ('Padres Estudiantes ITDB', itdb_count,               NARANJA_F),
        ('Familias Externas',       externos_count,            LAVANDA_F),
        ('TOTAL PADRES/EXTERNOS',   itdb_count + externos_count,
         PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')),
    ]:
        ws.append([label, val])
        r = ws.max_row
        ws.row_dimensions[r].height = 15
        for col in [1, 2]:
            c = ws.cell(r, col)
            c.fill = fill_obj; c.border = BORDER
            c.font = Font(bold=label.startswith('TOTAL'), size=10)
            c.alignment = Alignment(horizontal='center' if col == 2 else 'left', vertical='center')

    # Anchos de columna
    for i, w in enumerate([4, 12, 18, 18, 24, 28, 22, 13, 10, 14, 28, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="PADRES_EXTERNOS_' + str(gestion) + '.xlsx"'
    wb.save(response)
    return response


@login_required
def exportar_pdf_tutores(request):
    if not PISA_AVAILABLE:
        return HttpResponse("Instale xhtml2pdf.")
    gestion = date.today().year
    tutores = list(TutorPadre.objects.all().order_by('apellido_paterno', 'apellido_materno', 'nombres'))
    itdb = sum(1 for t in tutores if 'ITDB' in (t.ocupacion or '').upper() or 'ESTUDIANTE' in (t.ocupacion or '').upper())
    stats = {'itdb': itdb, 'externos': len(tutores) - itdb, 'total': len(tutores)}
    html = render_to_string('tutores/reporte_pdf.html', {
        'tutores': tutores, 'stats': stats, 'gestion': gestion
    })
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="TUTORES_{gestion}.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response

# =================== PRE-INSCRIPCION ===================
def preinscripcion_publica(request):
    form = PreInscripcionForm(request.POST or None, request.FILES or None)
    success = False
    if request.method == 'POST' and form.is_valid():
        form.save()
        success = True
        form = PreInscripcionForm()
    return render(request, 'beneficiarios/preinscripcion_publica.html', {'form': form, 'success': success})

def preinscripcion_confirmacion(request):
    return render(request, 'beneficiarios/preinscripcion_confirmacion.html')

@login_required
def lista_preinscripciones(request):
    if not _es_admin_director(request.user):
        messages.error(request, 'Solo Director o Administrador puede gestionar pre-inscripciones.')
        return redirect('dashboard')
    estado = request.GET.get('estado', 'pendiente')
    q = request.GET.get('q', '')
    preinscripciones = PreInscripcion.objects.filter(estado=estado).order_by('apellido_paterno_nino', 'nombres_nino')
    if q:
        preinscripciones = preinscripciones.filter(
            Q(apellido_paterno_nino__icontains=q) | Q(nombres_nino__icontains=q) |
            Q(CI_tutor__icontains=q) | Q(nombres_tutor__icontains=q)
        )
    return render(request, 'beneficiarios/lista_preinscripciones.html', {
        'preinscripciones': preinscripciones,
        'estado_filtro': estado,
        'pendientes_count': PreInscripcion.objects.filter(estado='pendiente').count(),
        'q': q,
    })

@login_required
def resolver_preinscripcion(request, pk):
    if not _es_admin_director(request.user):
        messages.error(request, 'Sin permiso.')
        return redirect('dashboard')
    p = get_object_or_404(PreInscripcion, pk=pk)
    if request.method == 'POST':
        accion = request.POST.get('accion')
        obs = request.POST.get('observaciones', '')
        if accion in ['aceptada', 'rechazada']:
            p.estado = accion
            p.observaciones_resolucion = obs
            p.resuelto_por = request.user
            p.fecha_resolucion = timezone.now()
            p.save()
            messages.success(request, f'Pre-inscripción {accion}.')
            return redirect('beneficiarios:lista_preinscripciones')
    return render(request, 'beneficiarios/detalle_preinscripcion.html', {'p': p})

@login_required
def eliminar_preinscripcion(request, pk):
    if not _es_admin_director(request.user):
        messages.error(request, 'Sin permiso.')
        return redirect('beneficiarios:lista_preinscripciones')
    p = get_object_or_404(PreInscripcion, pk=pk)
    nombre_nino = f"{p.nombres_nino} {p.apellido_paterno_nino}"
    if request.method == 'POST':
        p.delete()
        messages.success(request, f'Solicitud de {nombre_nino} eliminada.')
        return redirect('beneficiarios:lista_preinscripciones')
    return render(request, 'confirmar_eliminar.html', {
        'objeto': p, 'tipo': 'pre-inscripción', 'nombre_display': nombre_nino
    })

@login_required
def mover_preinscripcion(request, pk):
    if not _es_admin_director(request.user):
        messages.error(request, 'Sin permiso.')
        return redirect('beneficiarios:lista_preinscripciones')
    p = get_object_or_404(PreInscripcion, pk=pk)
    if p.estado != 'aceptada':
        messages.error(request, 'Solo se pueden mover pre-inscripciones aceptadas.')
        return redirect('beneficiarios:lista_preinscripciones')
    
    if request.method == 'POST':
        destino = request.POST.get('destino')
        if destino == 'beneficiario':
            return redirect('beneficiarios:crear_beneficiario_desde_preinscripcion', pk=pk)
        elif destino == 'tutor':
            return redirect('beneficiarios:crear_tutor_desde_preinscripcion', pk=pk)
    
    return render(request, 'beneficiarios/mover_preinscripcion.html', {'p': p})

@login_required
def crear_beneficiario_desde_preinscripcion(request, pk):
    if not _es_admin_director(request.user):
        messages.error(request, 'Sin permiso.')
        return redirect('beneficiarios:lista_preinscripciones')
    p = get_object_or_404(PreInscripcion, pk=pk)
    if p.estado != 'aceptada':
        messages.error(request, 'Solo se pueden mover pre-inscripciones aceptadas.')
        return redirect('beneficiarios:lista_preinscripciones')
    
    # Crear un nuevo Beneficiario con los datos del niño
    ci_beneficiario = request.POST.get('CI_beneficiario') if request.method == 'POST' else ''
    
    if request.method == 'POST':
        ci_beneficiario = request.POST.get('CI_beneficiario', '').strip()
        if not ci_beneficiario:
            messages.error(request, 'Debe proporcionar una cédula de identidad.')
            return render(request, 'beneficiarios/crear_beneficiario_desde_preinscripcion.html', {'p': p})
        
        # Verificar que la CI no exista ya
        if Beneficiario.objects.filter(CI_beneficiario=ci_beneficiario).exists():
            messages.error(request, f'Ya existe un beneficiario con la cédula {ci_beneficiario}.')
            return render(request, 'beneficiarios/crear_beneficiario_desde_preinscripcion.html', {'p': p})
        
        # Crear el Beneficiario
        beneficiario = Beneficiario(
            CI_beneficiario=ci_beneficiario,
            nombres=p.nombres_nino,
            apellido_paterno=p.apellido_paterno_nino,
            apellido_materno=p.apellido_materno_nino,
            contacto=p.telefono,
            foto=p.foto_nino,
            registrado_por=request.user
        )
        beneficiario.save()
        
        # Crear el registro del Niño con los datos de la pre-inscripción
        try:
            # Asegurar que apellido_paterno tenga un valor
            apellido_paterno = (p.apellido_paterno_nino or 'S/N').strip()
            if not apellido_paterno or apellido_paterno == 'S/N':
                apellido_paterno = 'S/N'
            
            # Copiar la foto si existe
            foto_nino = None
            if p.foto_nino:
                foto_nino = p.foto_nino
            
            nino = Nino(
                nombres=p.nombres_nino,
                apellido_paterno=apellido_paterno,
                apellido_materno=p.apellido_materno_nino or '',
                sexo=p.sexo_nino or 'M',
                fecha_nacimiento=p.fecha_nacimiento_nino,
                sala=p.sala_solicitada,
                foto=foto_nino,
                fecha_ingreso=date.today(),
                CI_beneficiario=beneficiario,
                registrado_por=request.user
            )
            nino.save()
            messages.success(request, f'Beneficiario {beneficiario.get_full_name()} y niño {nino.get_full_name()} creados exitosamente. Foto: {"Sí" if foto_nino else "No"}.')
        except Exception as e:
            import traceback
            messages.error(request, f'Error al crear el niño: {str(e)}. Contacte al administrador.')
            # Log para debugging
            print(f'ERROR al crear niño: {traceback.format_exc()}')
        
        return redirect('beneficiarios:lista_beneficiarios')
    
    return render(request, 'beneficiarios/crear_beneficiario_desde_preinscripcion.html', {'p': p})

@login_required
def crear_tutor_desde_preinscripcion(request, pk):
    if not _es_admin_director(request.user):
        messages.error(request, 'Sin permiso.')
        return redirect('beneficiarios:lista_preinscripciones')
    p = get_object_or_404(PreInscripcion, pk=pk)
    if p.estado != 'aceptada':
        messages.error(request, 'Solo se pueden mover pre-inscripciones aceptadas.')
        return redirect('beneficiarios:lista_preinscripciones')
    
    if request.method == 'POST':
        ci_tutor = request.POST.get('CI_tutor', '').strip()
        if not ci_tutor:
            messages.error(request, 'Debe proporcionar una cédula de identidad.')
            return render(request, 'beneficiarios/crear_tutor_desde_preinscripcion.html', {'p': p})
        
        # Verificar que la CI no exista ya
        if TutorPadre.objects.filter(CI_tutor=ci_tutor).exists():
            messages.error(request, f'Ya existe un tutor con la cédula {ci_tutor}.')
            return render(request, 'beneficiarios/crear_tutor_desde_preinscripcion.html', {'p': p})
        
        # Crear el TutorPadre
        tutor = TutorPadre(
            CI_tutor=ci_tutor,
            nombres=p.nombres_tutor,
            apellido_paterno=p.apellido_paterno_tutor,
            apellido_materno=p.apellido_materno_tutor,
            ocupacion=p.ocupacion,
            ocupacion_especifica=p.ocupacion_descripcion,
            carrera=p.carrera,
            anio_semestre=p.anio_semestre,
            turno=p.turno,
            contacto=p.telefono,
            registrado_por=request.user
        )
        tutor.save()

        # Intentar crear también el registro del niño y vincularlo al tutor
        try:
            # Asegurar que apellido_paterno tenga un valor
            apellido_paterno = (p.apellido_paterno_nino or 'S/N').strip()
            if not apellido_paterno or apellido_paterno == 'S/N':
                apellido_paterno = 'S/N'
            
            # Copiar la foto si existe
            foto_nino = None
            if p.foto_nino:
                foto_nino = p.foto_nino
            
            nino = Nino(
                nombres=p.nombres_nino,
                apellido_paterno=apellido_paterno,
                apellido_materno=p.apellido_materno_nino or '',
                sexo=p.sexo_nino or 'M',
                fecha_nacimiento=p.fecha_nacimiento_nino,
                sala=p.sala_solicitada,
                foto=foto_nino,
                fecha_ingreso=date.today(),
                CI_tutor_padre=tutor,
                registrado_por=request.user
            )
            nino.save()
            messages.success(request, f'Tutor {tutor.get_full_name()} y niño {nino.get_full_name()} creados exitosamente. Foto: {"Sí" if foto_nino else "No"}.')
        except Exception as e:
            import traceback
            messages.error(request, f'Error al crear el niño: {str(e)}. Tutor creado exitosamente.')
            # Log para debugging
            print(f'ERROR al crear niño: {traceback.format_exc()}')

        return redirect('beneficiarios:lista_tutores')

    return render(request, 'beneficiarios/crear_tutor_desde_preinscripcion.html', {'p': p})

@login_required
def eliminar_multiples_beneficiarios(request):
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            from ninos.models import Nino
            ninos_eliminados = Nino.objects.filter(CI_beneficiario__CI_beneficiario__in=ids).delete()
            Beneficiario.objects.filter(CI_beneficiario__in=ids).delete()
            messages.success(request, f'{len(ids)} padre(s)/IDS eliminado(s) junto con sus niños asociados.')
        else:
            messages.warning(request, 'No seleccionaste ninguno.')
    return redirect('beneficiarios:lista_beneficiarios')

@login_required
def eliminar_multiples_tutores(request):
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            from ninos.models import Nino
            Nino.objects.filter(CI_tutor_padre__CI_tutor__in=ids).delete()
            TutorPadre.objects.filter(CI_tutor__in=ids).delete()
            messages.success(request, f'{len(ids)} tutor(es) eliminado(s) junto con sus niños asociados.')
        else:
            messages.warning(request, 'No seleccionaste ninguno.')
    return redirect('beneficiarios:lista_tutores')