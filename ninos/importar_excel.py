"""
Importación Excel OFPROBOL 2025/2026
Columnas (datos desde fila 5):
 1=Nº  2=AP_PAT  3=AP_MAT  4=NOMBRES  5=SEXO  6=FEC_NAC  7=AÑO  8=MESES
 9=CI_NINO  10=DIRECCIÓN  11=PESO  12=TALLA  13=VACUNAS  14=SALA
 15=NOMBRE_TUTOR  16=CI_TUTOR  17=NRO_ITDB(vacío)  18=OCUPACION  19=CARRERA
 20=AÑO_SEM  21=TURNO  22=EDAD  23=CELULAR
"""
import openpyxl, re
from datetime import date, datetime
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from beneficiarios.models import Beneficiario, TutorPadre, CARRERA_CHOICES, TURNO_CHOICES
from ninos.models import Nino
from grupos.models import Grupo

SALA_MAP = {
    'LACTANTES':   'LACTANTES',
    'LACTANTE':    'LACTANTES',
    'INFANTE I':   'INFANTE I',
    'INFANTE II A':'INFANTE II A',
    'INFANTE IIA': 'INFANTE II A',
    'INFANTE II B':'INFANTE II B',
    'INFANTE IIB': 'INFANTE II B',
    'INFANTE II':  'INFANTE II A',
}

CARRERA_MAP = {
    'CONTADURIA GENERAL':'CONTADURIA_GENERAL','CONTADURÍA GENERAL':'CONTADURIA_GENERAL',
    'CONTADURA GENERAL':'CONTADURIA_GENERAL','CONTADIRIA GENERAL':'CONTADURIA_GENERAL',
    'CONTADURIA GENRAL':'CONTADURIA_GENERAL',
    'ADMINISTRACION DE EMPRESAS':'ADMINISTRACION_EMPRESAS',
    'SISTEMAS COMPUTACIONALES':'SISTEMAS_COMPUTACIONALES',
    'SISTEMAS INFORMATICOS':'SISTEMAS_COMPUTACIONALES',
    'EDUCACION':'EDUCACION','ENFERMERIA':'ENFERMERIA',
    'INGENIERIA CIVIL':'INGENIERIA_CIVIL',
    'INGENIERIA INDUSTRIAL':'INGENIERIA_INDUSTRIAL',
    'ELECTRICIDAD INDUSTRIAL':'INGENIERIA_INDUSTRIAL',
    'PSICOLOGIA':'PSICOLOGIA','DERECHO':'DERECHO',
    'INGLES':'INGLES','SECRETARIADO':'SISTEMAS_COMPUTACIONALES',
    'SECRETARIADO EJECUTIVO':'SISTEMAS_COMPUTACIONALES',
    'ARTES GRAFICAS':'SISTEMAS_COMPUTACIONALES',
    'DIRECCIÓN ACADEMICA ITDB':'OTRO',
}
TURNO_MAP = {'MAÑANA':'MANANA','MANANA':'MANANA','TARDE':'TARDE','NOCHE':'NOCHE'}

def _l(v): return '' if v is None else str(v).strip()
def _n(v): return _l(v).upper()

def _sala(raw):
    """Devuelve el nombre normalizado de sala desde el Excel."""
    n = _n(raw)
    # Buscar match exacto primero
    if n in SALA_MAP:
        return SALA_MAP[n]
    # Buscar match parcial
    for k, v in SALA_MAP.items():
        if k in n:
            return v
    return None

def _get_grupo(sala_normalizada):
    """Busca el Grupo en BD que corresponde a la sala normalizada.
    Si no existe, lo crea automáticamente."""
    if not sala_normalizada:
        return None

    # Cache en memoria para esta importación
    if not hasattr(_get_grupo, '_cache'):
        _get_grupo._cache = {}

    if sala_normalizada in _get_grupo._cache:
        return _get_grupo._cache[sala_normalizada]

    # Buscar en BD con match flexible
    grupo = None
    sala_up = sala_normalizada.upper()
    for g in Grupo.objects.all():
        g_up = g.nombre_grupo.upper().strip()
        if g_up == sala_up:
            grupo = g
            break
        # Match parcial para variaciones
        if 'LACTANTE' in sala_up and 'LACTANTE' in g_up:
            grupo = g; break
        if 'INFANTE II A' in sala_up and ('INFANTE II A' in g_up or 'INFANTE IIA' in g_up):
            grupo = g; break
        if 'INFANTE II B' in sala_up and ('INFANTE II B' in g_up or 'INFANTE IIB' in g_up):
            grupo = g; break
        if sala_up == 'INFANTE I' and g_up == 'INFANTE I':
            grupo = g; break

    # Si no existe, crearlo
    if not grupo:
        grupo = Grupo.objects.create(nombre_grupo=sala_normalizada)

    _get_grupo._cache[sala_normalizada] = grupo
    return grupo

def _carrera(raw):
    n = _n(raw)
    if n in CARRERA_MAP: return CARRERA_MAP[n]
    for k, v in CARRERA_MAP.items():
        if k in n: return v
    return 'OTRO'

def _turno(raw): return TURNO_MAP.get(_n(raw), None)

def _es_itdb(ocupacion):
    """Detecta si el padre es estudiante ITDB.
    El Excel usa 'ESTUDIANTE ITDB' (T antes de D) y también 'ESTIDIANTE ITDB'.
    """
    n = _n(ocupacion)
    return 'ITDB' in n

def _fecha(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    s = _l(val)
    for fmt in ('%d/%m/%Y','%Y-%m-%d','%d-%m-%Y','%d/%m/%y'):
        try: return datetime.strptime(s, fmt).date()
        except: pass
    m = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})', s)
    if m:
        try:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if y < 100: y += 2000
            return date(y, mo, d)
        except: pass
    return None

def _tel(val):
    if val is None: return ''
    return re.sub(r'\D', '', str(val))[:15]

def _peso(val):
    if val is None: return None
    s = str(val).strip().upper()
    if s in ('', 'FALTA', 'S/D', 'N/D', '-', 'FALTAAAAA'): return None
    s2 = re.sub(r'[KILOGRAMOSKG ]', '', s).replace(',', '.')
    m = re.search(r'(\d+\.?\d*)', s2)
    if m:
        try: return float(m.group(1))
        except: pass
    return None

def _talla(val):
    if val is None: return None
    s = str(val).strip().upper()
    if s in ('', 'FALTA', 'S/D', 'N/D', '-', 'FALTAAAAA'): return None
    s2 = s.replace(',', '.').replace('CM', '').strip()
    m = re.search(r'(\d+\.?\d*)', s2)
    if m:
        try: return float(m.group(1))
        except: pass
    return None

def _vacunas(val):
    return _n(val) in ('SI', 'SÍ', 'S', 'YES', 'Y', '1', 'TRUE', 'COMPLETAS')

def _partes(nombre):
    partes = _n(nombre).split()
    if not partes: return ('SIN NOMBRE', '', '')
    if partes[0] in ('TUTORA', 'TUTOR', 'DR', 'DRA', 'LIC'): partes = partes[1:]
    if not partes: return ('SIN NOMBRE', '', '')
    if len(partes) == 1: return (partes[0], '', '')
    if len(partes) == 2: return (partes[0], partes[1], '')
    # Nombre puede ser compuesto: tomar últimas 2 palabras como apellidos
    return (' '.join(partes[:-2]), partes[-2], partes[-1])

def _ci_temp(noms, ap_pat, tel):
    ini = ''
    for p in [ap_pat, noms]:
        if p: ini += p[0]
    t = tel[:6] if tel else '000000'
    return f"IMP-{ini}{t}".upper()[:15]


def procesar_excel(filepath, usuario):
    # Limpiar cache de grupos al inicio de cada importación
    if hasattr(_get_grupo, '_cache'):
        del _get_grupo._cache

    try:
        wb = openpyxl.load_workbook(filepath)
    except Exception as e:
        return None, None, f"Error al abrir: {e}"

    ws = wb.active
    resultados = []
    stats = {'creados': 0, 'existentes': 0, 'errores': 0, 'tutores_creados': 0}

    # Auto-detect fila de inicio (primera fila con datos reales, no encabezados)
    fila_inicio = 5
    for ri in range(1, 10):
        v = ws.cell(ri, 4).value
        if v and str(v).strip() and 'NOMBRE' not in str(v).upper():
            fila_inicio = ri
            break

    for row_idx, row in enumerate(ws.iter_rows(min_row=fila_inicio, values_only=True), start=fila_inicio):
        if all(v is None for v in row): continue
        row_list = list(row) + [None] * 5
        (nro, ap_pat_n, ap_mat_n, nombres_n, sexo_r,
         fec_nac_r, anio_r, meses_r,
         ci_nino_r, dir_r, peso_r, talla_r, vac_r, sala_r,
         nombre_tutor, ci_tutor_r, nro_itdb_r, ocup_r,
         carrera_r, anio_sem_r, turno_r, edad_r, cel_r) = row_list[:23]

        if not nombres_n and not ap_pat_n: continue

        sala_normalizada = _sala(sala_r)
        fi = {
            'fila': row_idx,
            'nino': f"{_n(ap_pat_n)} {_n(ap_mat_n)} {_n(nombres_n)}".strip(),
            'tutor': _n(nombre_tutor),
            'sala': _n(sala_r) if sala_r else '',
            'estado': '', 'mensaje': '', 'es_itdb': False, 'advertencias': [],
        }

        try:
            if not nombres_n: raise ValueError("Falta nombre del niño")

            fecha_nac = _fecha(fec_nac_r)
            sexo = _n(sexo_r)
            if sexo not in ('M', 'F'): sexo = 'M'

            peso   = _peso(peso_r)
            talla  = _talla(talla_r)
            vacunas = _vacunas(vac_r)
            direccion = _l(dir_r)
            ci_nino = _l(ci_nino_r)
            if ci_nino.upper() in ('FALTAAAAA', 'FALTA', 'S/D', '', '-'): ci_nino = ''

            tel = _tel(cel_r)
            ci_tutor = _l(ci_tutor_r)
            if ci_tutor.upper() in ('FALTAAAAA', 'FALTA', '', '-'): ci_tutor = ''

            es_itdb = _es_itdb(ocup_r)
            fi['es_itdb'] = es_itdb

            # Obtener grupo de la BD (o crearlo si no existe)
            grupo_obj = _get_grupo(sala_normalizada)

            tutor_obj = None
            benef_obj = None

            if es_itdb:
                # ── PADRE/ITDB ─────────────────────────────────────────
                carrera_k = _carrera(carrera_r)
                turno_k   = _turno(turno_r)
                ocup_esp  = _n(ocup_r)

                b = None
                if ci_tutor and len(ci_tutor) > 3:
                    b = Beneficiario.objects.filter(CI_beneficiario=ci_tutor).first()
                if not b and tel:
                    b = Beneficiario.objects.filter(contacto=tel).first()

                if b:
                    benef_obj = b
                    fi['advertencias'].append(f"Padre/ITDB ya existía: {b.get_full_name()}")
                else:
                    noms, ap_p, ap_m = _partes(nombre_tutor or '')
                    ci_k = ci_tutor if (ci_tutor and len(ci_tutor) > 3) else _ci_temp(noms, ap_p, tel)
                    # Si el CI ya existe, agregar sufijo para evitar duplicado
                    if Beneficiario.objects.filter(CI_beneficiario=ci_k).exists():
                        ci_k = (ci_k[:12] + str(row_idx))[:15]
                    benef_obj = Beneficiario(
                        CI_beneficiario=ci_k,
                        nombres=noms or 'SIN NOMBRE',
                        apellido_paterno=ap_p or 'S/N',
                        apellido_materno=ap_m,
                        ocupacion='ESTUDIANTE_ITDB',
                        ocupacion_especifica=ocup_esp[:100],
                        carrera=carrera_k,
                        anio_semestre=_n(anio_sem_r)[:50] if anio_sem_r else '',
                        turno=turno_k,
                        contacto=tel,
                        direccion=direccion or None,
                        registrado_por=usuario,
                    )
                    benef_obj.save()
                    stats['tutores_creados'] += 1
                    fi['advertencias'].append(f"Padre/ITDB creado: {noms} {ap_p} (CI={ci_k})")

            else:
                # ── PADRE/EXTERNO ──────────────────────────────────────
                t = None
                if ci_tutor and len(ci_tutor) > 3:
                    t = TutorPadre.objects.filter(CI_tutor=ci_tutor).first()
                if not t and tel:
                    t = TutorPadre.objects.filter(contacto=tel).first()

                if t:
                    tutor_obj = t
                    fi['advertencias'].append(f"Padre/Externo ya existía: {t.get_full_name()}")
                else:
                    noms, ap_p, ap_m = _partes(nombre_tutor or '')
                    ci_k = ci_tutor if (ci_tutor and len(ci_tutor) > 3) else _ci_temp(noms, ap_p, tel)
                    if TutorPadre.objects.filter(CI_tutor=ci_k).exists():
                        ci_k = (ci_k[:12] + str(row_idx))[:15]
                    tutor_obj = TutorPadre(
                        CI_tutor=ci_k,
                        nombres=noms or 'SIN NOMBRE',
                        apellido_paterno=ap_p or 'S/N',
                        apellido_materno=ap_m,
                        ocupacion='OTRO',
                        ocupacion_especifica=_n(ocup_r)[:100],
                        contacto=tel,
                        direccion=direccion or None,
                        registrado_por=usuario,
                    )
                    tutor_obj.save()
                    stats['tutores_creados'] += 1
                    fi['advertencias'].append(f"Padre/Externo creado: {noms} {ap_p} (CI={ci_k})")

            # ── Verificar duplicado niño ───────────────────────────────
            existe = Nino.objects.filter(
                nombres__iexact=_n(nombres_n),
                apellido_paterno__iexact=_n(ap_pat_n or ''),
            ).first()

            if existe:
                fi['estado']  = 'existente'
                fi['mensaje'] = f"Ya registrado (ID {existe.pk})"
                stats['existentes'] += 1
            else:
                nino = Nino(
                    nombres=_n(nombres_n),
                    apellido_paterno=_n(ap_pat_n or ''),
                    apellido_materno=_n(ap_mat_n or ''),
                    ci_nino=ci_nino or None,
                    sexo=sexo,
                    fecha_nacimiento=fecha_nac or date.today(),
                    fecha_ingreso=date.today(),
                    direccion=direccion or None,
                    sala=sala_normalizada,
                    id_grupo=grupo_obj,
                    peso_kg=peso,
                    talla_cm=talla,
                    vacunas_al_dia=vacunas,
                    CI_beneficiario=benef_obj,
                    CI_tutor_padre=tutor_obj,
                    registrado_por=usuario,
                )
                nino.save()
                fi['estado']  = 'creado'
                fi['mensaje'] = f"Creado en sala: {sala_normalizada or 'Sin sala'}"
                stats['creados'] += 1

        except Exception as e:
            fi['estado']  = 'error'
            fi['mensaje'] = str(e)
            stats['errores'] += 1

        resultados.append(fi)

    return resultados, stats, None


@login_required
def importar_excel_ninos(request):
    ctx = {'titulo': 'Importar Excel OFPROBOL'}
    if request.method == 'POST':
        archivo = request.FILES.get('archivo_excel')
        if not archivo:
            messages.error(request, 'Selecciona un archivo Excel (.xlsx)')
            return render(request, 'ninos/importar_excel.html', ctx)
        if not archivo.name.endswith('.xlsx'):
            messages.error(request, 'Solo se aceptan archivos .xlsx')
            return render(request, 'ninos/importar_excel.html', ctx)
        import tempfile, os
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            for chunk in archivo.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name
        try:
            resultados, stats, error = procesar_excel(tmp_path, request.user)
        finally:
            os.unlink(tmp_path)
        if error:
            messages.error(request, error)
            return render(request, 'ninos/importar_excel.html', ctx)
        ctx.update({'resultados': resultados, 'stats': stats, 'procesado': True})
        return render(request, 'ninos/importar_excel.html', ctx)
    return render(request, 'ninos/importar_excel.html', ctx)
