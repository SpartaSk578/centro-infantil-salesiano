import unicodedata
import openpyxl
from datetime import date
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Value
from django.db.models.functions import Upper, Replace
from django.http import HttpResponse
from django.template.loader import render_to_string
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from grupos.models import Grupo
from .models import Nino, DocumentoNino
from .forms import NinoForm, DocumentoNinoForm
from .importar_excel import importar_excel_ninos  # noqa: F401

try:
    from xhtml2pdf import pisa
    PISA_AVAILABLE = True
except ImportError:
    PISA_AVAILABLE = False

# ─── COLORES POR SALA ─────────────────────────────────────────────────────────
SALA_COLORS = {
    'LACTANTES':    {'fill': 'E2EFDA', 'hex': '#4CAF50', 'text': '#166534', 'label': 'Lactantes'},
    'LACTANTE':     {'fill': 'E2EFDA', 'hex': '#4CAF50', 'text': '#166534', 'label': 'Lactantes'},
    'INFANTE I':    {'fill': 'F1F5F9', 'hex': '#64748B', 'text': '#334155', 'label': 'Infante I'},
    'INFANTE II A': {'fill': 'FCE4D6', 'hex': '#EA580C', 'text': '#9A3412', 'label': 'Infante II A'},
    'INFANTE II B': {'fill': 'DBEAFE', 'hex': '#2563EB', 'text': '#1E3A8A', 'label': 'Infante II B'},
}
SALA_ORDER = ['LACTANTES', 'LACTANTE', 'INFANTE I', 'INFANTE II A', 'INFANTE II B', None]
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def get_sala_key(nombre_sala):
    if not nombre_sala: return None
    n = str(nombre_sala).upper()
    if 'LACTANTE' in n: return 'LACTANTES'
    if 'INFANTE II A' in n or 'INFANTE IIA' in n: return 'INFANTE II A'
    if 'INFANTE II B' in n or 'INFANTE IIB' in n: return 'INFANTE II B'
    if 'INFANTE I' in n: return 'INFANTE I'
    return None

def eliminar_tildes(cadena):
    if not cadena: return ""
    return ''.join(c for c in unicodedata.normalize('NFD', cadena) if unicodedata.category(c) != 'Mn')

def normalizar_queryset(queryset, campo):
    return queryset.annotate(**{f'{campo}_clean': Replace(
        Replace(Replace(Replace(Replace(Upper(campo),
        Value('Á'), Value('A')), Value('É'), Value('E')),
        Value('Í'), Value('I')), Value('Ó'), Value('O')),
        Value('Ú'), Value('U'))})

def limpiar(obj, attr):
    if not obj: return ""
    return str(getattr(obj, attr, '') or "").upper()

def get_ocupacion_full(obj):
    if not obj: return ""
    base = obj.get_ocupacion_display() if hasattr(obj, 'get_ocupacion_display') else ""
    esp = getattr(obj, 'ocupacion_especifica', '')
    return f"{base} ({esp})".upper() if esp else base.upper()

def get_celular(tutor):
    if not tutor: return ""
    for campo in ['contacto', 'celular', 'telefono']:
        val = getattr(tutor, campo, None)
        if val: return str(val)
    return ""

def sort_ninos_por_sala(ninos_list):
    def key(n):
        k = get_sala_key(str(n.id_grupo) if n.id_grupo else '')
        try: return (SALA_ORDER.index(k), n.apellido_paterno or '', n.nombres or '')
        except ValueError: return (99, n.apellido_paterno or '', n.nombres or '')
    return sorted(ninos_list, key=key)

def calc_stats(ninos_list):
    masc = sum(1 for n in ninos_list if n.sexo == 'M')
    fem  = sum(1 for n in ninos_list if n.sexo == 'F')
    itdb = externos = 0
    for n in ninos_list:
        t = n.responsable
        if t:
            ocu = getattr(t, 'ocupacion', '') or ''
            if 'ITDB' in str(ocu).upper() or 'ESTUDIANTE' in str(ocu).upper():
                itdb += 1
            else:
                externos += 1
        else:
            externos += 1
    return {'masc': masc, 'fem': fem, 'total': len(ninos_list), 'itdb': itdb, 'externos': externos}

# ─── VISTAS ──────────────────────────────────────────────────────────────────
@login_required
def lista_ninos(request):
    q = request.GET.get('q', '').strip()
    sala_id = request.GET.get('sala', '')
    ninos = Nino.objects.select_related('CI_beneficiario', 'id_grupo')
    if q:
        q_limpio = eliminar_tildes(q).upper()
        ninos = normalizar_queryset(ninos, 'nombres')
        ninos = normalizar_queryset(ninos, 'apellido_paterno')
        ninos = normalizar_queryset(ninos, 'apellido_materno')
        ninos = ninos.filter(
            Q(nombres_clean__icontains=q_limpio) |
            Q(apellido_paterno_clean__icontains=q_limpio) |
            Q(apellido_materno_clean__icontains=q_limpio)
        )
    if sala_id:
        ninos = ninos.filter(id_grupo_id=sala_id)
    ninos = ninos.order_by('apellido_paterno', 'apellido_materno', 'nombres')
    grupos = Grupo.objects.all()
    return render(request, 'ninos/lista.html', {'ninos': ninos, 'q': q, 'sala': sala_id, 'grupos': grupos})

@login_required
def detalle_nino(request, pk):
    nino = get_object_or_404(Nino, pk=pk)
    ead_items = [
        ('Motricidad Gruesa', nino.ead_motricidad_gruesa),
        ('Motricidad Fina', nino.ead_motricidad_fina),
        ('Audición y Lenguaje', nino.ead_audicion_lenguaje),
        ('Personal y Social', nino.ead_personal_social),
    ]
    return render(request, 'ninos/detalle.html', {
        'nino': nino,
        'asistencias': nino.asistencias.order_by('-fecha')[:10],
        'evaluaciones': nino.evaluaciones.order_by('-fecha')[:5],
        'ead_items': ead_items,
    })

@login_required
def crear_nino(request):
    if request.method == 'POST':
        form = NinoForm(request.POST, request.FILES)
        if form.is_valid():
            nino = form.save(commit=False)
            nino.registrado_por = request.user
            nino.save()
            messages.success(request, 'Niño registrado correctamente.')
            return redirect('ninos:lista_ninos')
    else:
        form = NinoForm()
    return render(request, 'ninos/form.html', {'form': form, 'titulo': 'Registrar Nuevo Niño'})

@login_required
def editar_nino(request, pk):
    nino = get_object_or_404(Nino, pk=pk)
    if request.method == 'POST':
        form = NinoForm(request.POST, request.FILES, instance=nino)
        if form.is_valid():
            form.save()
            messages.success(request, 'Información actualizada.')
            return redirect('ninos:detalle_nino', pk=pk)
    else:
        form = NinoForm(instance=nino)
    return render(request, 'ninos/form.html', {'form': form, 'titulo': 'Editar Información', 'objeto': nino})

@login_required
def subir_documento(request, pk):
    nino = get_object_or_404(Nino, pk=pk)
    if request.method == 'POST':
        form = DocumentoNinoForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.nino = nino
            doc.save()
            messages.success(request, 'Documento subido correctamente.')
            return redirect('ninos:detalle_nino', pk=pk)
    else:
        form = DocumentoNinoForm()
    return render(request, 'ninos/subir_documento.html', {'form': form, 'nino': nino})

@login_required
def eliminar_documento(request, pk):
    doc = get_object_or_404(DocumentoNino, pk=pk)
    nino_pk = doc.nino.pk
    if request.method == 'POST':
        doc.archivo.delete(save=False)
        doc.delete()
        messages.success(request, 'Documento eliminado.')
    return redirect('ninos:detalle_nino', pk=nino_pk)

def _es_admin_director(user):
    if not user.is_authenticated or not user.id_rol: return False
    return user.id_rol.nombre_rol.lower() in ['administrador', 'director', 'directora']

@login_required
def eliminar_nino(request, pk):
    if not _es_admin_director(request.user):
        messages.error(request, 'No tienes permiso.')
        return redirect('ninos:lista_ninos')
    nino = get_object_or_404(Nino, pk=pk)
    if request.method == 'POST':
        nino.delete()
        messages.success(request, 'Registro eliminado.')
        return redirect('ninos:lista_ninos')
    return render(request, 'ninos/confirmar_eliminacion.html', {'nino': nino})

@login_required
def eliminar_multiples_ninos(request):
    if request.method == 'POST':
        ids = request.POST.getlist('ids')
        if ids:
            from django.db import transaction
            with transaction.atomic():
                Nino.objects.filter(pk__in=ids).delete()
            messages.success(request, f'{len(ids)} niño(s) eliminado(s).')
        else:
            messages.warning(request, 'No seleccionaste ningún niño.')
    return redirect('ninos:lista_ninos')

# ─── EXCEL ────────────────────────────────────────────────────────────────────
@login_required
def exportar_excel_ninos(request):
    ninos_list = sort_ninos_por_sala(
        list(Nino.objects.all().select_related('CI_beneficiario', 'CI_tutor_padre', 'id_grupo'))
    )
    gestion = date.today().year
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inscritos"

    AZUL    = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    AMARILLO= PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    SALA_ROW_FILL = {
        'LACTANTES':    PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
        'INFANTE I':    PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid'),
        'INFANTE II A': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
        'INFANTE II B': PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid'),
    }

    # Título
    ws.row_dimensions[1].height = 28
    for i in range(1, 20):
        c = ws.cell(1, i)
        c.fill = AZUL
        c.font = Font(bold=True, size=11, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(1,1).value = f"CENTRO INFANTIL DON BOSQUITO — LISTA DE NIÑOS INSCRITOS — GESTIÓN {gestion}"

    # Headers
    headers = ['Nro','AP. PATERNO','AP. MATERNO','NOMBRES','SEXO','F. NACIMIENTO',
               'AÑO','MESES','SALA','NOMBRE TUTOR','OCUPACIÓN','CARRERA',
               'AÑO/SEM','TURNO','CELULAR','PESO(kg)','TALLA(cm)','E. NUTRICIONAL','VACUNAS']
    ws.append(headers)
    ws.row_dimensions[2].height = 40
    for cell in ws[2]:
        cell.fill = AMARILLO
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

    idx = 1
    for nino in ninos_list:
        sala_key = get_sala_key(str(nino.id_grupo) if nino.id_grupo else '')
        tutor = nino.responsable
        row = [
            idx,
            (nino.apellido_paterno or '').upper(),
            (nino.apellido_materno or '').upper(),
            (nino.nombres or '').upper(),
            nino.sexo,
            nino.fecha_nacimiento.strftime('%d/%m/%Y') if nino.fecha_nacimiento else '',
            nino.get_edad(),
            nino.get_meses_restantes(),
            str(nino.id_grupo).upper() if nino.id_grupo else 'SIN SALA',
            nino.get_responsable_name().upper(),
            get_ocupacion_full(tutor),
            limpiar(tutor, 'carrera'),
            limpiar(tutor, 'anio_semestre'),
            limpiar(tutor, 'turno'),
            get_celular(tutor),
            float(nino.peso_kg) if nino.peso_kg else '',
            float(nino.talla_cm) if nino.talla_cm else '',
            nino.get_estado_nutricional_display() if nino.estado_nutricional else '',
            'Sí' if nino.vacunas_al_dia else 'No',
        ]
        ws.append(row)
        r = ws.max_row
        row_fill = SALA_ROW_FILL.get(sala_key)
        ws.row_dimensions[r].height = 15
        for ci, cell in enumerate(ws[r], 1):
            if row_fill: cell.fill = row_fill
            cell.border = BORDER
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical='center')
        idx += 1

    # Anchos
    col_widths = [4,16,16,22,5,14,5,6,14,44,42,25,15,10,14,10,10,20,8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="LISTA_NINOS_{gestion}.xlsx"'
    wb.save(response)
    return response

@login_required
def exportar_pdf_ninos(request):
    if not PISA_AVAILABLE:
        return HttpResponse("Instale xhtml2pdf.")
    ninos_list = sort_ninos_por_sala(
        list(Nino.objects.all().select_related('CI_beneficiario', 'CI_tutor_padre', 'id_grupo'))
    )
    stats = calc_stats(ninos_list)
    html_string = render_to_string('ninos/reporte_pdf.html', {
        'ninos': ninos_list,
        'sala_colors': SALA_COLORS,
        'get_sala_key': get_sala_key,
        'stats': stats,
        'gestion': date.today().year,
    })
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="LISTA_NINOS_{date.today().year}.pdf"'
    pisa.CreatePDF(html_string, dest=response)
    return response
